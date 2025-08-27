#!/usr/bin/env python3
"""
Export Excel to a single JSON array.

Modes:
  - kv:    Two-column sheet where A=key, B=value -> [{"key", "value"}]
  - matrix: First row headers (from B), first column row keys (A) ->
            [{"row", "key", "value"}] across all non-empty cells

Usage:
  python scripts/export_sheet_json.py INPUT_XLSX [OUTPUT_JSON] [--mode kv|matrix]
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import load_workbook
except Exception:
    sys.stderr.write("Missing dependency 'openpyxl'. Install with: pip install openpyxl\n")
    raise


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export Excel to a single JSON array")
    parser.add_argument("input_xlsx", type=Path, help="Path to .xlsx file")
    parser.add_argument("output_json", type=Path, nargs="?", default=Path("sheet.json"))
    parser.add_argument("--sheet-name", dest="sheet_name", default=None)
    parser.add_argument("--mode", choices=["kv", "matrix"], default="matrix")
    parser.add_argument(
        "--matrix-output",
        choices=["triples", "records"],
        default="triples",
        help=(
            "For matrix mode: 'triples' -> [{row,key,value}], 'records' -> array of row objects keyed by headers."
        ),
    )
    parser.add_argument("--skip-header", action="store_true", help="For kv mode: skip first row")
    return parser.parse_args()


def load_worksheet(workbook_path: Path, sheet_name: Optional[str]):
    wb = load_workbook(filename=str(workbook_path), read_only=True, data_only=True)
    if sheet_name is None:
        return wb.worksheets[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
    return wb[sheet_name]


def export_kv(worksheet, skip_header: bool) -> List[Dict[str, Any]]:
    start_row = 2 if skip_header else 1
    result: List[Dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=start_row, values_only=True):
        key_cell = row[0] if len(row) >= 1 else None
        value_cell = row[1] if len(row) >= 2 else None
        if key_cell is None:
            continue
        result.append({"key": str(key_cell), "value": value_cell})
    return result


def export_matrix(worksheet) -> List[Dict[str, Any]]:
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers: Dict[int, str] = {}
    for idx, hdr in enumerate(header_row):
        if idx == 0:
            continue
        if hdr is None:
            continue
        headers[idx] = str(hdr)

    result: List[Dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_key = row[0] if len(row) >= 1 else None
        if row_key is None:
            continue
        row_key_str = str(row_key)
        for col_idx, header_name in headers.items():
            value = row[col_idx] if col_idx < len(row) else None
            result.append({"row": row_key_str, "key": header_name, "value": value})
    return result


def export_matrix_records(worksheet) -> List[Dict[str, Any]]:
    """Return an array of row objects keyed by headers from the first row."""
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers: Dict[int, str] = {}
    max_idx = 0
    for idx, hdr in enumerate(header_row):
        if hdr is None:
            # Allow None headers; skip such columns entirely
            continue
        headers[idx] = str(hdr)
        max_idx = max(max_idx, idx)

    records: List[Dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        # Build object for all known headers
        obj: Dict[str, Any] = {}
        for idx, header_name in headers.items():
            value = row[idx] if idx < len(row) else None
            obj[header_name] = value
        # Skip completely empty rows
        if any(v is not None and v != "" for v in obj.values()):
            records.append(obj)
    return records


def main() -> int:
    args = parse_args()
    if not args.input_xlsx.exists():
        sys.stderr.write(f"Input not found: {args.input_xlsx}\n")
        return 1

    ws = load_worksheet(args.input_xlsx, args.sheet_name)
    if args.mode == "kv":
        data = export_kv(ws, args.skip_header)
    else:
        if args.matrix_output == "records":
            data = export_matrix_records(ws)
        else:
            data = export_matrix(ws)

    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    with args.output_json.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


