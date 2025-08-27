#!/usr/bin/env python3
"""
Generate sheet_english.json by converting Arabic header names to English keys.

Behavior:
  - Reads the first worksheet by default.
  - Treats row 1 as headers and rows 2..N as data.
  - Builds an array of row objects, mapped using AR_TO_EN header map below.
  - Only mapped columns are included in the output.

Usage:
  python scripts/export_sheet_english.py INPUT_XLSX [OUTPUT_JSON]

Options:
  --sheet-name <name>  Optional worksheet name (defaults to first sheet)
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import load_workbook
except Exception:
    sys.stderr.write("Missing dependency 'openpyxl'. Install with: pip install openpyxl\n")
    raise


# Map Arabic header names -> desired English keys
# Add aliases to be resilient to minor header text changes.
AR_TO_EN: Dict[str, str] = {
    # Core fields seen in the sheet
    "رقم أولي": "preliminary_number",
    "رقم اولي": "preliminary_number",
    "موضوع التشريع": "legislation_subject",
    "رقم التشريع": "legislation_number",
    "تاريخ التشريع": "legislation_date_hijri",
    "تاريخ رفع التشريع": "legislation_submission_date",
    "ATTACHMENT(كود المرفق على التشريع)": "attachment_id",
    "ATTACHMENT (كود المرفق على التشريع)": "attachment_id",
    "كود الركيزة": "pillar_code",
    "اسم الركيزة": "pillar_name",
    "اسم الركيزة الأم": "main_pillar_name",
    "اسم الركيزة الإم": "main_pillar_name",  # common typo/variant
    # Authorization
    "كود التفويض": "authorization_id",
    "اسم التفويض": "authorization_name",
    "الصلاحية": "authorization_id",
    "اسم الصلاحية": "authorization_name",
    # Classification
    "كود التصنيف": "classification_id",
    "اسم التصنيف": "classification_name",
    "التصنيف": "classification_name",
    # Unit
    "كود الوحدة": "unit_code",
    "اسم الوحدة": "unit_name",
    # Additional attachment header variant
    "(كود المرفق على التشريع)ATTACHMENT_ID": "attachment_id",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert sheet headers to English keys.")
    parser.add_argument("input_xlsx", type=Path, help="Path to input .xlsx file")
    parser.add_argument("output_json", type=Path, nargs="?", default=Path("sheet_english.json"))
    parser.add_argument("--sheet-name", dest="sheet_name", default=None)
    parser.add_argument(
        "--list-unmapped",
        action="store_true",
        help="List unmapped headers from the first row and exit",
    )
    return parser.parse_args()


def load_worksheet(workbook_path: Path, sheet_name: Optional[str]):
    wb = load_workbook(filename=str(workbook_path), read_only=True, data_only=True)
    if sheet_name is None:
        return wb.worksheets[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
    return wb[sheet_name]


def build_header_index(header_row: List[Any]) -> Dict[int, str]:
    """Return mapping of column index -> English key using AR_TO_EN; skip unmapped headers."""
    header_map: Dict[int, str] = {}
    for idx, header_value in enumerate(header_row):
        if header_value is None:
            continue
        header_text = str(header_value).strip()
        english_key = AR_TO_EN.get(header_text)
        if english_key:
            header_map[idx] = english_key
    return header_map


def to_serializable(value: Any) -> Any:
    # openpyxl already gives Python primitives. Just return as-is.
    return value


def export_records(workbook_path: Path, sheet_name: Optional[str]) -> List[Dict[str, Any]]:
    ws = load_worksheet(workbook_path, sheet_name)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_index = build_header_index(list(header_row))

    records: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        obj: Dict[str, Any] = {}
        for col_idx, english_key in header_index.items():
            value = row[col_idx] if col_idx < len(row) else None
            obj[english_key] = to_serializable(value)
        if any(v is not None and v != "" for v in obj.values()):
            records.append(obj)
    return records


def main() -> int:
    args = parse_args()
    if not args.input_xlsx.exists():
        sys.stderr.write(f"Input not found: {args.input_xlsx}\n")
        return 1

    try:
        ws = load_worksheet(args.input_xlsx, args.sheet_name)
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header_values = [str(v).strip() for v in header_row if v is not None]
        if args.list_unmapped:
            unmapped = [h for h in header_values if h not in AR_TO_EN]
            sys.stdout.write(json.dumps(unmapped, ensure_ascii=False, indent=2) + "\n")
            return 0
        # normal export path
        data = export_records(args.input_xlsx, args.sheet_name)
    except Exception as exc:
        sys.stderr.write(f"Failed converting sheet: {exc}\n")
        return 1

    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    with args.output_json.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


